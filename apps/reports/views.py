"""
API views for WorkReport listing, detail, and export.
"""
from datetime import date, timedelta

from django.db.models import Count, Q
from django.http import HttpResponse
from rest_framework import generics
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from apps.accounts.permissions import apply_report_access_filter
from apps.stats.models import WorkEntry
from .models import WorkReport
from .serializers import ReportListSerializer, ReportDetailSerializer

# ---- Shared queryset builder ----

def _build_report_queryset(request):
    """Build base WorkReport queryset from request query params.

    Shared by ReportListView and ReportExportView to avoid duplicated
    filter / search / permission logic.
    """
    qs = WorkReport.objects.select_related("creator", "department") \
        .annotate(entry_count=Count("work_entries", distinct=True))

    # Exclude demo seed data
    qs = qs.exclude(dingtalk_report_id__startswith="demo_report_")

    date_from = request.query_params.get("date_from")
    date_to = request.query_params.get("date_to")
    if date_from:
        qs = qs.filter(report_date__gte=date_from)
    if date_to:
        qs = qs.filter(report_date__lte=date_to)
    if not date_from and not date_to:
        qs = qs.filter(report_date__gte=date.today() - timedelta(days=30))

    username = request.query_params.get("username")
    if username:
        qs = qs.filter(creator__username=username)

    department = request.query_params.get("department")
    if department:
        qs = qs.filter(department__name=department)

    search = request.query_params.get("search")
    if search:
        qs = qs.filter(
            Q(contents__field_value__icontains=search)
        ).distinct()

    # Role-based access filter
    qs = apply_report_access_filter(qs, request.user)
    return qs


# ---- Pagination ----

class ReportListPagination(PageNumberPagination):
    page_size = 15
    page_size_query_param = "page_size"
    max_page_size = 50
    ordering = "-report_date"


# ---- Views ----

class ReportListView(generics.ListAPIView):
    """
    GET /api/reports/
    Query params: ?page=1&page_size=15&date_from=2026-06-01&date_to=2026-06-30
                  &username=admin&department=技术部&search=项目A
    """
    serializer_class = ReportListSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = ReportListPagination

    def get_queryset(self):
        return _build_report_queryset(self.request)


class ReportDetailView(generics.RetrieveAPIView):
    """
    GET /api/reports/{id}/
    """
    serializer_class = ReportDetailSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = WorkReport.objects.select_related("creator", "department") \
            .prefetch_related("contents", "work_entries__project", "work_entries__employee")
        # Exclude demo seed data
        qs = qs.exclude(dingtalk_report_id__startswith="demo_report_")
        return apply_report_access_filter(qs, self.request.user)


# ---- Excel export styling ----

_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGN = Alignment(vertical="center", wrap_text=False)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Column index → (header, width)
_SUMMARY_COLS = [
    ("序号", 6),
    ("日志日期", 12),
    ("员工姓名", 12),
    ("用户名", 14),
    ("部门", 16),
    ("状态", 8),
    ("工时条目数", 10),
    ("入库时间", 20),
]

_ENTRY_COLS = [
    ("序号", 6),
    ("日志日期", 12),
    ("员工", 12),
    ("部门", 16),
    ("项目名称", 22),
    ("项目代码", 14),
    ("工时(h)", 10),
    ("工作类型", 10),
    ("任务描述", 54),
    ("工作日期", 12),
    ("置信度", 8),
]


def _write_header(ws, cols):
    """Write styled header row for the given column definitions."""
    for col_idx, (label, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_cell(ws, row, col, value, align=None):
    """Write a data cell with standard border + optional alignment."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = _THIN_BORDER
    if align:
        cell.alignment = align
    return cell


class ReportExportView(generics.GenericAPIView):
    """
    GET /api/reports/export/
    Query params: same as list (date_from, date_to, username, department, search)
                  but WITHOUT pagination — exports ALL matching records.

    Returns an .xlsx file with two sheets:
      1. 日志汇总 — one row per report
      2. 工时明细 — one row per work entry
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        reports = _build_report_queryset(request).order_by("-report_date")

        # 单独导出模式：只导出指定 ID 的报告
        report_ids = request.query_params.getlist("report_ids")
        if report_ids:
            ids = [int(i) for i in report_ids if i.isdigit()]
            if ids:
                reports = reports.filter(id__in=ids)

        # Fetch linked work entries in one query (avoid N+1)
        entry_qs = WorkEntry.objects.filter(report__in=reports) \
            .select_related("employee", "department", "project", "report") \
            .order_by("report__report_date", "employee__username")

        # Build a lookup: report_id → [entries]
        entries_by_report = {}
        for e in entry_qs:
            entries_by_report.setdefault(e.report_id, []).append(e)

        wb = Workbook()

        # ---- Sheet 1: 日志汇总 ----
        ws1 = wb.active
        ws1.title = "日志汇总"
        _write_header(ws1, _SUMMARY_COLS)
        ws1.freeze_panes = "A2"

        row = 2
        for r in reports:
            entry_count = entries_by_report.get(r.id, [])
            status_display = "已提交" if r.status == "submitted" else "草稿"
            created_str = r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else ""
            vals = [
                row - 1,
                str(r.report_date) if r.report_date else "",
                r.creator.get_full_name() or r.creator.username,
                r.creator.username,
                r.department.name if r.department else "",
                status_display,
                len(entry_count),
                created_str,
            ]
            for col_idx, val in enumerate(vals, 1):
                _write_cell(ws1, row, col_idx, val)
            row += 1

        # ---- Sheet 2: 工时明细 ----
        ws2 = wb.create_sheet("工时明细")
        _write_header(ws2, _ENTRY_COLS)
        ws2.freeze_panes = "A2"

        row = 2
        for r in reports:
            for e in entries_by_report.get(r.id, []):
                vals = [
                    row - 1,
                    str(r.report_date) if r.report_date else "",
                    e.employee.get_full_name() or e.employee.username if e.employee else "",
                    e.department.name if e.department else "",
                    e.project.name if e.project else "未归类",
                    e.project.code if e.project else "",
                    float(e.hours),
                    e.get_task_type_display(),
                    e.task_description or "",
                    str(e.date) if e.date else "",
                    e.confidence,
                ]
                for col_idx, val in enumerate(vals, 1):
                    _write_cell(ws2, row, col_idx, val)
                row += 1

        # Response
        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"work_reports_{date.today()}.xlsx"
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(resp)
        return resp
